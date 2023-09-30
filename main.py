from openpyxl import Workbook, load_workbook
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from uuid import uuid4

STORAGE_FILE_NAME = 'data.xlsx'

class Person(BaseModel):
    firstName: str
    lastName: str
    age: int

class PersonWithId(Person):
    id: str
   
#создаем экземпляр класса FastAPI который является веб сервером   
app = FastAPI() 

wb = None
try:
    wb = load_workbook(filename=STORAGE_FILE_NAME)
except:
  wb = Workbook()
  wb.save(filename=STORAGE_FILE_NAME)

persons_sheet = None
try:
    persons_sheet = wb.get_sheet_by_name("Person")
except:
    persons_sheet = wb.create_sheet("Person")
    wb.save(filename=STORAGE_FILE_NAME)

# #откр эксель фай и создает экз КЛАССА workbook
# wb = load_workbook(filename='data.xlsx')
# # sheet_ranges = wb.cr("Person")
# # sheet_ranges = wb.get_sheet_by_name("Person")
# persons_sheet = wb.active
# # print(sheet_ranges['D18'].value)

# allow cors requests
origins = [
    "*",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

@app.get("/")
def read_root():
    return {"Hello": "World"}

# Read
@app.post("/user")
def create_user(user: Person):
    id = str(uuid4())
    ws = persons_sheet
    ws.append([id, user.firstName, user.lastName, user.age])
    wb.save(filename=STORAGE_FILE_NAME)
    pid = PersonWithId(id=id, firstName=user.firstName, lastName=user.lastName, age=user.age)
    return pid

# Read
@app.get("/users")
def get_users():
    persons = []
    for row in persons_sheet.iter_rows(values_only=True):
        if row[0]:
            p = PersonWithId(id=row[0], firstName=row[1], lastName=row[2], age=row[3])
            persons.append(p)
    return persons

# update
@app.put("/user")
def update_users(user: PersonWithId):
    print('Updating user')
    for row in persons_sheet.iter_rows():
        cell = row[0]
        if cell.value == user.id:
            persons_sheet['B' + str(cell.row)] = user.firstName
            persons_sheet['C' + str(cell.row)] = user.lastName
            persons_sheet['D' + str(cell.row)] = user.age
            wb.save(filename=STORAGE_FILE_NAME)
            return user
    raise Exception('User not found')

# delete
@app.delete("/user")
def delete_user(id: str | None = None):
    if id == None:
        raise Exception("No ID provided")
    
    print('DELETE USER:' + id)
    user = None
    for row in persons_sheet.iter_rows():
        cell = row[0]
        if cell.value == id:
            idx = cell.row
            print('Found index of id: ' + id + ' \nindex:' + str(idx))
            user = PersonWithId(id=row[0].value, firstName=row[1].value, lastName=row[2].value, age=row[3].value)
            persons_sheet.delete_rows(idx)
    return user
    

